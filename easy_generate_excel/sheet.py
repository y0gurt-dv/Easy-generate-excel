from copy import copy
from typing import Literal, Union

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_interval, get_column_letter

ALL = Literal['all']


class Sheet:
    def __init__(
            self,
            name: str,
            headers: list,
            data: list[list],
            bold_header: bool = True,
            auto_filter: bool = True,
            center_cols_indexes: Union[ALL, list[int], None] = 'all',
            center_headers_indexes: Union[ALL, list[int], None] = 'all',
            not_center_cols_indexes: Union[ALL, list[int], None] = None,
            not_center_headers_indexes: Union[ALL, list[int], None] = None,
            cell_expansion: Union[int, float] = 1.2,
            min_cell_width: int = 9):
        """Sheet config

        :param name: Sheet name
        :type name: str

        :param headers: Headers list
        :type headers: list

        :param data: Data list. Data will be read by data[row_idx][col_idx]
        :type data: list

        :param bold_header: Need bold header
        :type bold_header: bool

        :param auto_filter: Need enable auto filter in headers
        :type auto_filter: bool

        :param center_cols_indexes: Cols indexes of data cols which will be center
        :type center_cols_indexes: list

        :param center_headers_indexes: Cols indexes of headers cols which will be center
        :type center_headers_indexes: list

        :param not_center_cols_indexes: Cols indexes of data cols which will ``NOT`` be center
        :type not_center_cols_indexes: list

        :param not_center_headers_indexes: Cols indexes of headers cols which will ``NOT`` be center
        :type not_center_headers_indexes: list

        :param cell_expansion: The percentage by which the cell width will be increased
        :type cell_expansion: int or float

        :param min_cell_width: Min cell width (Excel default 9)
        :type min_cell_width: int

        """
        self.name = name
        self.headers = headers
        self.data = data
        self.auto_filter = auto_filter
        self.bold_header = bold_header
        self.cell_expansion = cell_expansion
        self.min_cell_width = min_cell_width

        if center_cols_indexes is None:
            self.center_cols_indexes = []
        else:
            self.center_cols_indexes = center_cols_indexes

        if center_headers_indexes is None:
            self.center_headers_indexes = []
        else:
            self.center_headers_indexes = center_headers_indexes

        if not_center_cols_indexes is None:
            self.not_center_cols_indexes = []
        else:
            self.not_center_cols_indexes = not_center_cols_indexes

        if not_center_headers_indexes is None:
            self.not_center_headers_indexes = []
        else:
            self.not_center_headers_indexes = not_center_headers_indexes

    @property
    def max_row(self):
        """
            Return max row index (count data with headers)
        """
        return len(self.data) + 1

    @property
    def max_col(self):
        """
            Return max column index (count headers)
        """
        return len(self.headers)

    @property
    def cols_interval(self):
        """
            Return list of all columns
        """
        return get_column_interval(1, self.max_col)

    @property
    def full_data(self):
        """
            Return data with headers in index 0
        """
        data = list(self.data)
        data.insert(0, self.headers)
        return data

    def __repr__(self) -> str:
        return f'<Sheet name:{self.name}>'

    def _need_center_col(self, row_idx: int, col_idx: int) -> bool:
        """
            Need center cell by coords (row, col)
        """
        if row_idx == 0:
            indexes_list = self.center_headers_indexes
            not_indexes_list = self.not_center_headers_indexes
        else:
            indexes_list = self.center_cols_indexes
            not_indexes_list = self.not_center_headers_indexes

        if indexes_list == 'all':
            return True

        if ((col_idx in not_indexes_list or not_indexes_list == 'all')
                and col_idx not in indexes_list):
            return False

        return col_idx in indexes_list

    def _get_cell_width(self, base_width: int) -> int:
        """
            Return new cell width
        """
        return max(
            round(base_width * self.cell_expansion),
            self.min_cell_width
        )

    @classmethod
    def from_dict(cls, raw: dict):
        """
            Return new obj from dict or json
        """

        new_instance = cls(
            name=raw['name'],
            headers=raw['headers'],
            data=raw['data'],
            auto_filter=raw.get('auto_filter', True),
            bold_header=raw.get('auto_filter', True),
            center_cols_indexes=raw.get('center_cols_indexes', 'all'),
            center_headers_indexes=raw.get('center_headers_indexes', 'all'),
            not_center_cols_indexes=raw.get('not_center_cols_indexes'),
            not_center_headers_indexes=raw.get('not_center_headers_indexes'),
            cell_expansion=raw.get('cell_expansion', 1.2),
            min_cell_width=raw.get('min_cell_width', 9),
        )
        return new_instance

    def create(self, workbook: Workbook):
        """
            Create new worksheet in workbook
        """
        worksheet = workbook.create_sheet(self.name)
        max_rows_widths = {idx: 0 for idx in self.cols_interval}
        data = self.full_data
        cells: list[list[Cell]] = worksheet.iter_rows(
            min_row=1,
            min_col=1,
            max_row=self.max_row,
            max_col=self.max_col
        )

        for row in cells:
            for cell in row:
                row_idx = cell.row - 1
                col_idx = cell.column - 1
                col_letter = cell.column_letter

                value = str(data[row_idx][col_idx])
                width = len(max(value.split('\n'), key=len))

                cell.value = value
                max_rows_widths[col_letter] = max(
                    max_rows_widths[col_letter],
                    width
                )

                if row_idx == 0 and self.bold_header:
                    new_font = copy(cell.font)
                    new_font.bold = True
                    cell.font = new_font

                if self._need_center_col(row_idx, col_idx):
                    cell.alignment = Alignment(
                        vertical='center',
                        horizontal='center'
                    )

        for letter, width in max_rows_widths.items():
            worksheet.column_dimensions[letter].width = self._get_cell_width(
                base_width=width
            )

        if self.auto_filter:
            worksheet.auto_filter.ref = 'A1:{}1'.format(
                get_column_letter(self.max_col)
            )
